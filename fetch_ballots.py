from openpyxl import load_workbook
from os import listdir
import pickle

ballots = {}

for file in listdir("data"):
    sheet = load_workbook(filename=f"data/{file}").active
    mayor_col = 0
    for row in sheet.iter_rows(max_row=1):
        for cell in row:
            if "DEM Mayor Choice 1" in cell.value:
                mayor_col = cell.column
    for row in sheet.iter_rows(min_row=2, min_col=mayor_col, max_col=mayor_col+4):
        ballot_tup = tuple((cell.value for cell in row))
        if ballot_tup in ballots:
            ballots[ballot_tup] += 1
        else:
            ballots[ballot_tup] = 1
    print(file)

with open("ballots.pickle", 'wb') as ballotsFile:
    pickle.dump(ballots, ballotsFile)