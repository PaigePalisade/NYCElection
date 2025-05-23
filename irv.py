from openpyxl import load_workbook
from os import listdir
import pickle

name_sheet = load_workbook(filename="2021P_CandidacyID_To_Name.xlsx").active

id2name = {}
index2id = [
    "217572",
    "219978",
    "219469",
    "218127",
    "218491",
    "217796",
    "217605",
    "221141",
    "221183",
    "218117",
    "218922",
    "217654",
    "221458"
]

n = len(index2id)
id2index = {k: v for (k,v) in zip(index2id, range(n))}

for row in name_sheet.iter_rows(min_row=2, max_col=2):
    id2name[str(row[0].value)] = row[1].value

with open("ballots.pickle", 'rb') as file:
    ballots = pickle.load(file)

first_round = [0 for i in range(n)]

for ballot in ballots:
    for i in range(5):
        if ballot[i] == '217572' or ballot[i] == '219978' or ballot[i] == 'overvote':
            if ballot[i] != 'overvote' and ballot[i] != 'Write-in':
                first_round[id2index[ballot[i]]] += ballots[ballot]
            break

for i in range(n):
    print(f"{id2name[index2id[i]]:20}{first_round[i]:10}",)

# for ballot in ballots:
#     if not 'overvote' in ballot:
#         for i in range(5):
#             if ballot[i] == 'undervote':
#                 break
#             if ballot[i] != 'undervote' and ballot[i] != 'overvote' and ballot[i] != 'Write-in':
#                 for j in range(n):
#                     if index2id[j] not in ballot[0:i+1]:
#                         voteMatrix[id2index[ballot[i]]][j] += ballots[ballot]
#                         # voteMatrix[j][id2index[ballot[i]]] -= ballots[ballot]


# print(" " * 21, end="")

# for i in range(n):
#     print(f"{id2name[index2id[i]]:>21}", end="")

# print()

# for r in range(n):
#     print(f"{id2name[index2id[r]]:20}", end="")
#     for e in voteMatrix[r]:
#         print(f"{e:21}",end="")
#     print()



