from openpyxl import load_workbook
from os import listdir
import pickle

name_sheet = load_workbook(filename="Primary Election 2025 - 06-24-2025_CandidacyID_To_Name.xlsx").active

id2name = {}
index2id = ['254365', '255950', '257465', '254130', '254607', '254393', '258719', '258821', '254286', '257441', '254052']

n = len(index2id)

pairwiseMatrix = [[0 for c in range(n)] for r in range(n)]

id2index = {k: v for (k,v) in zip(index2id, range(len(index2id)))}

for row in name_sheet.iter_rows(min_row=2, max_col=2):
    id2name[str(row[0].value)] = row[1].value

with open("ballots.pickle", 'rb') as file:
    ballots = pickle.load(file)

for ballot in ballots:
    for i in range(5):
        # ballot is discarded after the first instance of an overvote
        if ballot[i] == 'overvote':
            break
        # undervotes are skipped and we are not counting write-in's for the purpose of this tabulation. if candidate is ranked twice, the second instance is skipped
        if ballot[i] != 'undervote' and ballot[i] != 'Write-in' and not ballot[i] in ballot[0:i]:
            # for every other candidate
            for j in range(n):
                # if they haven't been ranked already
                if index2id[j] not in ballot[0:i+1]:
                    # add their vote to the pairwise matrix
                    pairwiseMatrix[id2index[ballot[i]]][j] += ballots[ballot]



# create a difference matrix
diffMatrix = [[pairwiseMatrix[r][c] - pairwiseMatrix[c][r] for c in range(n)] for r in range(n)]

# get candidate order
wins = []

for candidate in range(n):
    w = 0
    for other in range(n):
        if (diffMatrix[candidate][other] > 0):
            w += 1
    wins.append((candidate, w))

wins.sort(key=lambda x : -x[1])

# print tables
def print_matrix(m):
    print(" " * 20, end="")
    for c in wins:
        print(f"{id2name[index2id[c[0]]]:>21}", end="")

    print()

    for c in wins:
        print(f"{id2name[index2id[c[0]]]:20}", end="")
        for o in wins:
            print(f"{m[c[0]][o[0]]:21}",end="")
        print()
    print()

print_matrix(pairwiseMatrix)
print_matrix(diffMatrix)